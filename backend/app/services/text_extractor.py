from __future__ import annotations

import asyncio
import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple, TYPE_CHECKING

import pytesseract
from PIL import Image
import fitz  # PyMuPDF

from app.config import settings

if TYPE_CHECKING:
    pass  # avoid circular imports

logger = logging.getLogger(__name__)

# Global semaphore for heavy OCR/PDF operations
OCR_SEMAPHORE = asyncio.Semaphore(2)


class TextExtractorMixin:
    """Mixin: text extraction and OCR methods."""

    async def _stage_text_extraction(self, document, document_dir: Path, progress_callback=None):
        """Stage 2: Extract text from document using OCR if needed."""
        from app.models.schemas import OCRResult
        import json

        if "\x00" in (document.original_filename or ""):
            raise ValueError("Invalid filename (embedded null byte)")
        file_path = document_dir / "original" / document.original_filename
        text_dir = document_dir / "text"
        text_dir.mkdir(exist_ok=True)

        mime_type = document.mime_type
        pages = []
        combined_text = ""
        ocr_used = False

        try:
            if mime_type == "application/pdf":
                pages, combined_text, ocr_used = await self._extract_pdf_text(file_path, progress_callback=progress_callback, document_id=document.id)
            elif mime_type.startswith("image/"):
                pages, combined_text, ocr_used = await self._extract_image_text(file_path)
            elif mime_type in ["application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                              "application/msword"]:
                pages, combined_text, ocr_used = await self._extract_docx_text(file_path)
            elif mime_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              "application/vnd.ms-excel"]:
                pages, combined_text, ocr_used = await self._extract_xlsx_text(file_path)
            else:
                raise ValueError(f"Unsupported file type: {mime_type}")

        except Exception as e:
            logger.error(f"Text extraction failed: {e}")
            raise

        # Determine OCR quality
        ocr_quality = self._assess_ocr_quality(combined_text, ocr_used)

        result = OCRResult(
            pages=pages,
            combined_text=combined_text,
            ocr_used=ocr_used,
            ocr_quality=ocr_quality
        )

        # Save artifacts
        with open(text_dir / "extracted.json", "w", encoding="utf-8") as f:
            json.dump(result.dict(), f, indent=2, ensure_ascii=False)

        with open(text_dir / "extracted.txt", "w", encoding="utf-8") as f:
            f.write(combined_text)

        return result

    async def _ocr_with_rotation_detection(self, img: Image.Image) -> str:
        """Perform OCR on image, trying different rotations (0, 90, 180, 270) and return best result.

        Optimized: First tries 0°, only tries other rotations if 0° doesn't produce good results.

        Args:
            img: PIL Image to perform OCR on

        Returns:
            Best OCR text result from all rotations
        """
        results = []

        # First try 0° (no rotation) - most common case
        try:
            async with OCR_SEMAPHORE:
                text_0 = await asyncio.to_thread(
                    pytesseract.image_to_string, img, config=settings.tesseract_config
                )

            alnum_count_0 = sum(1 for c in text_0 if c.isalnum())
            word_count_0 = len(text_0.split())
            score_0 = alnum_count_0 * 2 + word_count_0

            results.append({
                'angle': 0,
                'text': text_0,
                'score': score_0,
                'alnum_count': alnum_count_0,
                'word_count': word_count_0
            })

            logger.debug(f"OCR rotation 0°: {alnum_count_0} alnum chars, {word_count_0} words, score: {score_0}")

            # If 0° produces good results (reasonable amount of text), skip other rotations
            # Threshold: at least 50 alphanumeric characters and 10 words indicates readable text
            if alnum_count_0 >= 50 and word_count_0 >= 10:
                logger.debug(f"OCR 0° produces good results ({alnum_count_0} alnum, {word_count_0} words), skipping other rotations")
                return text_0

        except Exception as e:
            logger.warning(f"OCR failed for rotation 0°: {e}")

        # If 0° didn't produce good results, try other rotations
        for angle in [90, 180, 270]:
            try:
                # Rotate image
                rotated_img = img.rotate(-angle, expand=True)  # Negative for counter-clockwise

                # Perform OCR (wrapped in semaphore + to_thread)
                async with OCR_SEMAPHORE:
                    text = await asyncio.to_thread(
                        pytesseract.image_to_string, rotated_img, config=settings.tesseract_config
                    )

                # Score the result: count alphanumeric characters and words
                alnum_count = sum(1 for c in text if c.isalnum())
                word_count = len(text.split())

                # Prefer results with more alphanumeric characters and words
                score = alnum_count * 2 + word_count

                results.append({
                    'angle': angle,
                    'text': text,
                    'score': score,
                    'alnum_count': alnum_count,
                    'word_count': word_count
                })

                logger.debug(f"OCR rotation {angle}°: {alnum_count} alnum chars, {word_count} words, score: {score}")

            except Exception as e:
                logger.warning(f"OCR failed for rotation {angle}°: {e}")
                continue

        if not results:
            logger.warning("All OCR rotations failed, returning empty string")
            return ""

        # Sort by score (descending) and return best result
        results.sort(key=lambda x: x['score'], reverse=True)
        best = results[0]

        if best['angle'] != 0:
            logger.info(f"Best OCR result found at {best['angle']}° rotation ({best['alnum_count']} alnum chars, {best['word_count']} words)")

        return best['text']

    async def _extract_pdf_text(self, file_path: Path, progress_callback=None, document_id: int = None) -> Tuple[List[Dict[str, Any]], str, bool]:
        """Extract text from PDF using multiple methods, falling back to OCR if needed."""
        pages = []
        combined_text = ""
        ocr_used = False

        try:
            doc = fitz.open(str(file_path))

            # Pre-extract pdfminer once for the whole document (not per page)
            pdfminer_page_texts: list[str] = []
            try:
                from pdfminer.high_level import extract_text as pdfminer_extract
                full_text = await asyncio.to_thread(pdfminer_extract, str(file_path))
                pdfminer_page_texts = full_text.split('\f')
            except Exception as e:
                logger.debug(f"pdfminer pre-extraction failed: {e}")

            # Pre-load pypdf reader once
            pypdf_reader = None
            try:
                from pypdf import PdfReader
                pypdf_reader = await asyncio.to_thread(PdfReader, str(file_path))
            except Exception as e:
                logger.debug(f"pypdf pre-load failed: {e}")

            for page_num in range(len(doc)):
                page = doc.load_page(page_num)
                text = ""
                source = "text-layer"

                # Method 1: PyMuPDF text extraction
                try:
                    text = page.get_text("text")
                    if not text or len(text.strip()) < 50:
                        text = page.get_text("rawtext")
                        source = "text-layer-raw"
                except Exception as e:
                    logger.debug(f"PyMuPDF text extraction failed for page {page_num}: {e}")

                # Method 2: pypdf fallback
                is_garbage = self._is_garbage_text(text)
                if len(text.strip()) < 200 or self._is_mostly_empty(text) or is_garbage:
                    logger.info(f"Page {page_num}: PyMuPDF inadequate (len={len(text.strip())}, garbage={is_garbage}), trying pypdf")
                    if pypdf_reader and page_num < len(pypdf_reader.pages):
                        try:
                            text = pypdf_reader.pages[page_num].extract_text() or ""
                            source = "pypdf"
                        except Exception as e:
                            logger.debug(f"pypdf page extraction failed for page {page_num}: {e}")

                    # Method 3: pdfminer fallback (already extracted above)
                    is_garbage = self._is_garbage_text(text)
                    if len(text.strip()) < 200 or self._is_mostly_empty(text) or is_garbage:
                        logger.info(f"Page {page_num}: pypdf inadequate, trying pdfminer")
                        try:
                            if page_num < len(pdfminer_page_texts):
                                text = pdfminer_page_texts[page_num]
                            elif pdfminer_page_texts:
                                text = pdfminer_page_texts[0]
                            source = "pdfminer"
                        except Exception as e:
                            logger.debug(f"pdfminer page lookup failed for page {page_num}: {e}")

                # Method 3: If all text extraction fails or text is garbage, use OCR
                is_garbage = self._is_garbage_text(text)
                needs_ocr = (
                    len(text.strip()) < 200 or
                    self._is_mostly_empty(text) or
                    is_garbage
                )
                if needs_ocr:
                    if len(text.strip()) < 200:
                        reason = "too short"
                    elif self._is_mostly_empty(text):
                        reason = "mostly empty"
                    else:
                        reason = "garbage text"
                    logger.info(f"Page {page_num}: All extractors failed ({reason}, len={len(text.strip())}, garbage={is_garbage}), using OCR")
                    async def render_page():
                        async with OCR_SEMAPHORE:
                            return await asyncio.to_thread(page.get_pixmap, dpi=150)
                    pix = await render_page()
                    img = Image.open(BytesIO(pix.tobytes("png")))
                    text = await self._ocr_with_rotation_detection(img)
                    source = "ocr"
                    ocr_used = True

                pages.append({
                    "page": page_num,
                    "source": source,
                    "text": text
                })
                combined_text += text + "\n\n"

                logger.info(f"Page {page_num}: Extracted {len(text.strip())} chars using {source}")

                # Per-page progress: 10% → 45% spread over pages
                if progress_callback and document_id and len(doc) > 0:
                    page_pct = int(10 + ((page_num + 1) / len(doc)) * 35)
                    try:
                        await progress_callback(document_id, page_pct, "extracting_text")
                    except Exception:
                        pass

            doc.close()

            logger.info(f"PDF extraction complete: {len(pages)} pages, {len(combined_text.strip())} total chars, OCR used: {ocr_used}")

        except Exception as e:
            logger.warning(f"PDF text extraction failed, trying full OCR: {e}")
            # Full OCR fallback with rotation detection
            try:
                doc = fitz.open(str(file_path))
                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    async def render_page():
                        async with OCR_SEMAPHORE:
                            return await asyncio.to_thread(page.get_pixmap, dpi=150)
                    pix = await render_page()
                    img = Image.open(BytesIO(pix.tobytes("png")))
                    text = await self._ocr_with_rotation_detection(img)

                    pages.append({
                        "page": page_num,
                        "source": "ocr",
                        "text": text
                    })
                    combined_text += text + "\n\n"
                    ocr_used = True

                doc.close()
            except Exception as ocr_error:
                logger.error(f"OCR fallback also failed: {ocr_error}")
                raise

        return pages, combined_text, ocr_used

    async def _extract_image_text(self, file_path: Path) -> Tuple[List[Dict[str, Any]], str, bool]:
        """Extract text from image using OCR with rotation detection."""
        img = Image.open(file_path)
        text = await self._ocr_with_rotation_detection(img)

        pages = [{
            "page": 0,
            "source": "ocr",
            "text": text
        }]

        return pages, text, True

    async def _extract_docx_text(self, file_path: Path) -> Tuple[List[Dict[str, Any]], str, bool]:
        """Extract text from DOCX file."""
        from docx import Document as DocxDocument
        doc = DocxDocument(file_path)
        text = ""

        for para in doc.paragraphs:
            text += para.text + "\n"

        pages = [{
            "page": 0,
            "source": "docx",
            "text": text
        }]

        return pages, text, False

    async def _extract_xlsx_text(self, file_path: Path) -> Tuple[List[Dict[str, Any]], str, bool]:
        """Extract text from XLSX file."""
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        text = ""

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"Sheet: {sheet_name}\n"

            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                text += row_text + "\n"

            text += "\n"

        pages = [{
            "page": 0,
            "source": "xlsx",
            "text": text
        }]

        return pages, text, False

    def _is_mostly_empty(self, text: str) -> bool:
        """Check if text is mostly empty or contains mostly symbols."""
        # Remove whitespace and check ratio of alphanumeric characters
        cleaned = re.sub(r'\s+', '', text)
        if len(cleaned) == 0:
            return True

        alpha_numeric = len(re.findall(r'[a-zA-Z0-9]', cleaned))
        return (alpha_numeric / len(cleaned)) < 0.1

    def _is_garbage_text(self, text: str) -> bool:
        """Detect garbage text from corrupted PDF encoding or wrong character mapping.

        This happens when PDF has embedded fonts with custom encoding that don't
        map to standard characters, resulting in readable-looking but meaningless text.
        """
        if not text or len(text.strip()) < 50:
            logger.info(f"Garbage detection: text too short ({len(text.strip()) if text else 0} chars)")
            return True

        # Clean text for analysis
        cleaned = text.strip()

        # Check 1: Too many special/punctuation characters relative to letters
        # Garbage text often has patterns like &'/!%.&'1&+$
        letters = len(re.findall(r'[a-zA-Z]', cleaned))
        special_chars = len(re.findall(r'[&%$#@!^*+=<>|\\~`\'\"/]', cleaned))
        if letters > 0 and special_chars / letters > 0.5:
            logger.info(f"Garbage detection: high special char ratio ({special_chars}/{letters} = {special_chars/letters:.2f})")
            return True

        # Check 2: Very few recognizable words (3+ consecutive letters)
        words = re.findall(r'[a-zA-Z]{3,}', cleaned)
        if len(cleaned) > 100 and len(words) < 5:
            logger.info(f"Garbage detection: too few words ({len(words)} in {len(cleaned)} chars)")
            return True

        # Check 3: High ratio of control characters or weird unicode
        control_chars = len(re.findall(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', cleaned))
        if control_chars > len(cleaned) * 0.05:
            logger.info(f"Garbage detection: high control char ratio ({control_chars}/{len(cleaned)})")
            return True

        # Check 4: Check for common Dutch/English word patterns
        # If text has enough letters but no common word patterns, it's likely garbage
        common_patterns = [
            r'\b(de|het|een|en|van|in|is|dat|op|te|voor|met|aan|zijn|worden|door|als|naar|uit|over|ook|meer|tot|bij|nog|dan|wel|om|of|kan|dit|niet|maar|er|al|wat|hebben|was|jaar|zou|gaan|na|zo|ons|die|hier|wordt)\b',  # Dutch
            r'\b(the|a|an|is|are|was|were|be|been|have|has|had|do|does|did|will|would|could|should|may|might|must|shall|can|to|of|and|in|for|on|with|at|by|from|or|as|but|not|this|that|it|he|she|we|they|you|i)\b',  # English
        ]

        text_lower = cleaned.lower()
        common_word_count = 0
        for pattern in common_patterns:
            common_word_count += len(re.findall(pattern, text_lower))

        # If we have significant text but very few common words, likely garbage
        if len(words) > 10 and common_word_count < 3:
            logger.info(f"Garbage detection: no common words found ({common_word_count} in {len(words)} words)")
            return True

        # Check 5: Look for specific garbage patterns common in corrupted PDFs
        # Patterns like #+0#,1#.-#. or &'/!%.&'1&+$
        garbage_patterns = [
            r'[#&][^a-zA-Z\s]{3,}',  # # or & followed by 3+ non-letter chars
            r'[+\-*/]{2,}',  # Multiple math operators in a row
            r'\d[#&%]\d',  # Number-symbol-number patterns
        ]
        garbage_matches = 0
        for pattern in garbage_patterns:
            garbage_matches += len(re.findall(pattern, cleaned))

        if garbage_matches >= 3:
            logger.info(f"Garbage detection: found {garbage_matches} garbage patterns")
            return True

        logger.debug(f"Garbage detection: text appears valid ({len(cleaned)} chars, {len(words)} words, {common_word_count} common)")
        return False

    def _assess_ocr_quality(self, text: str, ocr_used: bool) -> str:
        """Assess OCR quality based on text characteristics."""
        if not ocr_used:
            return "high"

        cleaned = re.sub(r'\s+', '', text)
        if len(cleaned) < 100:
            return "low"

        # Check for common OCR errors
        error_indicators = [
            r'[^\x00-\x7F]',  # Non-ASCII characters (could be OCR artifacts)
            r'\d{10,}',      # Very long numbers (potential OCR errors)
            r'[|@#$%^&*]{3,}',  # Multiple special characters together
        ]

        error_score = 0
        for pattern in error_indicators:
            error_score += len(re.findall(pattern, text))

        if error_score > len(cleaned) * 0.05:  # More than 5% potential errors
            return "low"
        elif error_score > len(cleaned) * 0.02:  # More than 2% potential errors
            return "medium"
        else:
            return "high"
